# -*- coding: utf-8 -*-
"""
Gera a tabela mestre de probabilidades normalizadas:
Kalshi + Polymarket + Oddschecker + media das tres fontes.

Este script salva apenas XLSX, com formatacao condicional em azul.
"""
import argparse
import unicodedata
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NOME_ARQUIVO_MESTRE = (
    "TABELA_MESTRA_probabilidades_normalizadas_"
    "Kalshi_Polymarket_Oddschecker_{data}.xlsx"
)

SINONIMOS = {
    "united states": "usa",
    "korea republic": "south korea",
    "czech republic": "czechia",
    "cote d'ivoire": "ivory coast",
    "cabo verde": "cape verde",
    "turkiye": "turkey",
    "ir iran": "iran",
    "dr congo": "congo dr",
    "curacao": "curacao",
    "bosnia and herzegovina": "bosnia",
}


def chave_selecao(nome):
    texto = str(nome).strip().lower()
    texto = "".join(
        c for c in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(c)
    )
    texto = texto.replace("’", "'").replace("`", "'")
    texto = " ".join(texto.split())
    return SINONIMOS.get(texto, texto)


def aplicar_formatacao(caminho):
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Comparativo"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    for col in range(2, 6):
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = "0.00%"
            ws[f"{letter}{row}"].alignment = Alignment(horizontal="right")

    blue_scale = ColorScaleRule(
        start_type="min", start_color="EFF6FF",
        mid_type="percentile", mid_value=50, mid_color="93C5FD",
        end_type="max", end_color="1D4ED8",
    )
    ws.conditional_formatting.add(f"B2:E{ws.max_row}", blue_scale)

    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 16

    wb.save(caminho)


def gerar_tabela(mercados_path, oddschecker_path):
    mercados = pd.read_excel(mercados_path)
    odds = pd.read_excel(oddschecker_path)

    mercados["team_key_join"] = mercados["Selecao"].map(chave_selecao)
    odds["team_key_join"] = odds["Selecao"].map(chave_selecao)

    comparativo = mercados[[
        "Selecao_PT",
        "team_key_join",
        "prob_kalshi_normalizada",
        "prob_polymarket_normalizada",
    ]].merge(
        odds[["team_key_join", "prob_implicita_media_normalizada"]],
        on="team_key_join",
        how="left",
        validate="one_to_one",
    )

    sem_odds = comparativo[
        comparativo["prob_implicita_media_normalizada"].isna()
    ]["Selecao_PT"].tolist()
    if sem_odds:
        raise RuntimeError(f"Selecoes sem Oddschecker: {sem_odds}")

    comparativo = comparativo.rename(columns={
        "Selecao_PT": "Selecao",
        "prob_kalshi_normalizada": "Kalshi",
        "prob_polymarket_normalizada": "Polymarket",
        "prob_implicita_media_normalizada": "Oddschecker",
    })
    comparativo["Media_3_fontes"] = comparativo[[
        "Kalshi",
        "Polymarket",
        "Oddschecker",
    ]].mean(axis=1)

    colunas = ["Selecao", "Kalshi", "Polymarket", "Oddschecker", "Media_3_fontes"]
    return comparativo[colunas].sort_values(
        "Media_3_fontes", ascending=False
    ).reset_index(drop=True)


def main():
    parser = argparse.ArgumentParser(
        description="Gera a tabela mestre comparativa entre Kalshi, Polymarket e Oddschecker."
    )
    parser.add_argument("--mercados", required=True, help="XLSX de Kalshi + Polymarket")
    parser.add_argument("--oddschecker", required=True, help="XLSX do Oddschecker")
    parser.add_argument("--data", default=date.today().isoformat(), help="Data AAAA-MM-DD")
    parser.add_argument("--saida-dir", default=None, help="Pasta onde salvar o XLSX final")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    analise_dir = script_dir.parent
    saida_dir = Path(args.saida_dir) if args.saida_dir else (
        analise_dir / "resultados" / args.data
    )
    saida_dir.mkdir(parents=True, exist_ok=True)
    saida = saida_dir / NOME_ARQUIVO_MESTRE.format(data=args.data)

    df = gerar_tabela(Path(args.mercados), Path(args.oddschecker))
    df.to_excel(saida, index=False)
    aplicar_formatacao(saida)

    print(f"Salvo em: {saida}")
    print(f"Linhas: {len(df)}")
    print(df.head(10).to_string(index=False))


if __name__ == "__main__":
    main()
