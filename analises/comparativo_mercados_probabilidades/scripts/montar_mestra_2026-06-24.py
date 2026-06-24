# -*- coding: utf-8 -*-
"""
Monta a TABELA_MESTRA da fase 03 (pos rodada-2, 2026-06-24).

Difere do gerar_tabela_mestra_comparativo.py padrao apenas no tratamento dos
times que sairam do mercado de "campeao":

  - Nesta rodada o Oddschecker parou de precificar 4 azaroes (Jordania, Tunisia,
    Haiti, Turquia) e o Polymarket parou de precificar 5 (esses 4 + Panama).
  - Oddschecker ausente -> imputado no PISO do Oddschecker (menor prob presente)
    e a coluna e renormalizada para somar 1 sobre as 48.
  - Polymarket ausente -> fica NaN; a Media_3_fontes usa a media das fontes
    disponiveis (skipna), depois renormalizada para somar 1.
  - Media_3_ajustada: C = min(Media_3_fontes) - 1/100000; subtrai C de todas e
    renormaliza (afina a cauda; mesma formula validada em 2026-06-18).
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
from gerar_tabela_mestra_comparativo import chave_selecao

OUT = SCRIPT_DIR.parent / "resultados" / "03_pos_rodada-2_2026-06-24"
MERCADOS = OUT / "mercados_predicao_2026-06-24.xlsx"
ODDS = OUT / "oddschecker_tabela_com_probs_2026-06-24.xlsx"
SAIDA = OUT / "TABELA_MESTRA_probabilidades_normalizadas_Kalshi_Polymarket_Oddschecker_2026-06-24.xlsx"


def main():
    mercados = pd.read_excel(MERCADOS)
    odds = pd.read_excel(ODDS)

    mercados["team_key_join"] = mercados["Selecao"].map(chave_selecao)
    odds["team_key_join"] = odds["Selecao"].map(chave_selecao)

    comp = mercados[[
        "Selecao_PT", "team_key_join",
        "prob_kalshi_normalizada", "prob_polymarket_normalizada",
    ]].merge(
        odds[["team_key_join", "prob_implicita_media_normalizada"]],
        on="team_key_join", how="left", validate="one_to_one",
    )

    comp = comp.rename(columns={
        "Selecao_PT": "Selecao",
        "prob_kalshi_normalizada": "Kalshi",
        "prob_polymarket_normalizada": "Polymarket",
        "prob_implicita_media_normalizada": "Oddschecker",
    })

    # Oddschecker ausente -> piso do proprio Oddschecker, depois renormaliza p/ somar 1.
    faltam = comp[comp["Oddschecker"].isna()]["Selecao"].tolist()
    piso = comp["Oddschecker"].min()
    print(f"Times sem Oddschecker (imputados no piso {piso:.6f}): {faltam}")
    comp["Oddschecker"] = comp["Oddschecker"].fillna(piso)
    comp["Oddschecker"] = comp["Oddschecker"] / comp["Oddschecker"].sum()

    # Media das 3 fontes (skipna p/ quem nao tem Polymarket), renormalizada.
    media = comp[["Kalshi", "Polymarket", "Oddschecker"]].mean(axis=1, skipna=True)
    comp["Media_3_fontes"] = media / media.sum()

    # Media_3_ajustada:
    #  1) zera as selecoes sem chance matematica na Copa;
    #  2) das positivas, pega o menor p_min e subtrai C = p_min - 0.0001 de TODAS
    #     as positivas (as zeradas ficam em 0), levando a menor para 0.0001;
    #  3) renormaliza a coluna para somar 1.
    SEM_CHANCE = ["Panamá", "Jordânia", "Tunísia", "Haiti", "Turquia"]
    nao_encontradas = [s for s in SEM_CHANCE if s not in set(comp["Selecao"])]
    if nao_encontradas:
        raise RuntimeError(f"Selecoes a zerar nao encontradas: {nao_encontradas}")
    positivas = ~comp["Selecao"].isin(SEM_CHANCE)
    print(f"Zeradas (sem chance): {SEM_CHANCE}")

    p_min = comp.loc[positivas, "Media_3_fontes"].min()
    C = p_min - 0.0001
    ajust = comp["Media_3_fontes"] - C
    ajust = ajust.where(positivas, 0.0)
    comp["Media_3_ajustada"] = ajust / ajust.sum()
    menor_pos = comp.loc[comp.loc[positivas, "Media_3_fontes"].idxmin(), "Selecao"]
    print(f"Ajuste da cauda: menor positivo={menor_pos} p_min={p_min:.6f}  C={C:.6f}")

    cols = ["Selecao", "Kalshi", "Polymarket", "Oddschecker",
            "Media_3_fontes", "Media_3_ajustada"]
    df = comp[cols].sort_values("Media_3_fontes", ascending=False).reset_index(drop=True)
    df.to_excel(SAIDA, index=False)
    formatar(SAIDA)

    print(f"\nSalvo: {SAIDA}")
    print("Somas:", {c: round(df[c].sum(), 4) for c in cols[1:]})
    pd.set_option("display.width", 200)
    print(df.head(12).to_string(index=False))
    print("...")
    print(df.tail(6).to_string(index=False))


def formatar(caminho):
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Comparativo"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)
    for col in range(2, 7):  # B..F
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = "0.00%"
            ws[f"{letter}{row}"].alignment = Alignment(horizontal="right")
    blue = ColorScaleRule(
        start_type="min", start_color="EFF6FF",
        mid_type="percentile", mid_value=50, mid_color="93C5FD",
        end_type="max", end_color="1D4ED8",
    )
    ws.conditional_formatting.add(f"B2:F{ws.max_row}", blue)
    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 16
    wb.save(caminho)


if __name__ == "__main__":
    main()
