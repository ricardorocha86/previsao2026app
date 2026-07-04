# -*- coding: utf-8 -*-
"""
Monta a TABELA_MESTRA da fase 05 (pos 16-avos, 2026-07-04).

Regras desta fase:
  - usa Kalshi + Polymarket + Oddschecker quando disponiveis;
  - se o arquivo Oddschecker existir mas nao listar todas as selecoes, imputa
    o piso do proprio Oddschecker e renormaliza a coluna;
  - se uma fonte inteira estiver ausente, Media_3_fontes usa as fontes
    disponiveis com skipna e renormaliza;
  - Media_3_ajustada zera as 32 selecoes eliminadas e renormaliza os 16
    vencedores dos 16-avos. Sem ajuste de cauda.
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
from gerar_tabela_mestra_comparativo import chave_selecao

OUT = SCRIPT_DIR.parent / "resultados" / "05_pos_16avos_2026-07-04"
MERCADOS = OUT / "mercados_predicao_2026-07-04.xlsx"
ODDS = OUT / "oddschecker_tabela_com_probs_2026-07-04.xlsx"
SAIDA = OUT / "TABELA_MESTRA_probabilidades_normalizadas_Kalshi_Polymarket_Oddschecker_2026-07-04.xlsx"

SEM_CHANCE = [
    "Arábia Saudita", "Catar", "Coreia do Sul", "Curaçau", "Escócia",
    "Haiti", "Iraque", "Irã", "Jordânia", "Nova Zelândia", "Panamá",
    "Tcheca", "Tunísia", "Turquia", "Uruguai", "Uzbequistão",
    "África do Sul", "Alemanha", "Holanda", "Japão", "Suécia",
    "Costa do Marfim", "Equador", "RD do Congo", "Bósnia e Herzegovina",
    "Senegal", "Croácia", "Áustria", "Argélia", "Cabo Verde", "Gana",
    "Austrália",
]


def carregar_oddschecker(comp):
    if not ODDS.exists():
        print(f"Oddschecker ausente: {ODDS}. Media_3_fontes usara as fontes disponiveis.")
        comp["Oddschecker"] = pd.NA
        return comp

    odds = pd.read_excel(ODDS)
    odds["team_key_join"] = odds["Selecao"].map(chave_selecao)
    comp = comp.merge(
        odds[["team_key_join", "prob_implicita_media_normalizada"]],
        on="team_key_join", how="left", validate="one_to_one",
    )
    comp = comp.rename(columns={
        "prob_implicita_media_normalizada": "Oddschecker",
    })

    faltam = comp[comp["Oddschecker"].isna()]["Selecao"].tolist()
    piso = comp["Oddschecker"].min()
    if pd.notna(piso):
        print(f"Times sem Oddschecker (imputados no piso {piso:.6f}): {faltam}")
        comp["Oddschecker"] = comp["Oddschecker"].fillna(piso)
        comp["Oddschecker"] = comp["Oddschecker"] / comp["Oddschecker"].sum()
    else:
        print("Oddschecker sem probabilidades validas. Media_3_fontes seguira sem essa fonte.")
        comp["Oddschecker"] = pd.NA
    return comp


def main():
    mercados = pd.read_excel(MERCADOS)
    mercados["team_key_join"] = mercados["Selecao"].map(chave_selecao)

    comp = mercados[[
        "Selecao_PT", "team_key_join",
        "prob_kalshi_normalizada", "prob_polymarket_normalizada",
    ]].rename(columns={
        "Selecao_PT": "Selecao",
        "prob_kalshi_normalizada": "Kalshi",
        "prob_polymarket_normalizada": "Polymarket",
    })
    comp = carregar_oddschecker(comp)

    fontes = ["Kalshi", "Polymarket", "Oddschecker"]
    media = comp[fontes].mean(axis=1, skipna=True)
    comp["Media_3_fontes"] = media / media.sum()

    nao_encontradas = [s for s in SEM_CHANCE if s not in set(comp["Selecao"])]
    if nao_encontradas:
        raise RuntimeError(f"Selecoes a zerar nao encontradas: {nao_encontradas}")

    positivas = ~comp["Selecao"].isin(SEM_CHANCE)
    ajust = comp["Media_3_fontes"].where(positivas, 0.0)
    comp["Media_3_ajustada"] = ajust / ajust.sum()
    print(f"Zeradas ({len(SEM_CHANCE)} eliminados): {SEM_CHANCE}")
    print(f"Positivas apos ajuste: {int((comp['Media_3_ajustada'] > 0).sum())}")

    cols = ["Selecao", "Kalshi", "Polymarket", "Oddschecker",
            "Media_3_fontes", "Media_3_ajustada"]
    df = comp[cols].sort_values("Media_3_fontes", ascending=False).reset_index(drop=True)
    df.to_excel(SAIDA, index=False)
    formatar(SAIDA)

    print(f"\nSalvo: {SAIDA}")
    print("Somas:", {c: round(float(df[c].sum(skipna=True)), 6) for c in cols[1:]})
    pd.set_option("display.width", 200)
    print(df.head(34).to_string(index=False))


def formatar(caminho):
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Comparativo"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)
    for col in range(2, 7):
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = "0.00%"
            ws[f"{letter}{row}"].alignment = Alignment(horizontal="right")
    blue = ColorScaleRule(
        start_type="min", start_color="EFF6FF",
        mid_type="percentile", mid_value=50, mid_color="93C5FD",
        end_type="max", end_color="1D4ED8",
    )
    ws.conditional_formatting.add(f"B2:F{ws.max_row}", blue)
    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 16
    wb.save(caminho)


if __name__ == "__main__":
    main()
