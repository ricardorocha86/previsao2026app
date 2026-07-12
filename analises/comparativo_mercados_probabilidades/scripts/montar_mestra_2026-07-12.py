# -*- coding: utf-8 -*-
"""Monta a TABELA_MESTRA da fase 07 (pós-quartas, 2026-07-12)."""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
from gerar_tabela_mestra_comparativo import chave_selecao

OUT = SCRIPT_DIR.parent / "resultados" / "07_pos_quartas_2026-07-12"
MERCADOS = OUT / "mercados_predicao_2026-07-12.xlsx"
ODDS = OUT / "oddschecker_tabela_com_probs_2026-07-12.xlsx"
SAIDA = OUT / "TABELA_MESTRA_probabilidades_normalizadas_Kalshi_Polymarket_Oddschecker_2026-07-12.xlsx"

SEM_CHANCE = [
    "Alemanha", "Argélia", "Arábia Saudita", "Austrália", "Bélgica", "Brasil",
    "Bósnia e Herzegovina", "Cabo Verde", "Canadá", "Catar", "Colômbia",
    "Coreia do Sul", "Costa do Marfim", "Croácia", "Curaçau", "Egito",
    "Equador", "Escócia", "Estados Unidos", "Gana", "Haiti", "Holanda",
    "Iraque", "Irã", "Japão", "Jordânia", "Marrocos", "México",
    "Noruega", "Nova Zelândia", "Panamá", "Paraguai", "Portugal", "RD do Congo",
    "Senegal", "Suécia", "Suíça", "Tcheca", "Tunísia", "Turquia", "Uruguai",
    "Uzbequistão", "África do Sul", "Áustria",
]


def carregar_oddschecker(comp):
    if not ODDS.exists():
        print(f"Oddschecker ausente: {ODDS}. Media_3_fontes usara as fontes disponiveis.")
        comp["Oddschecker"] = pd.NA
        return comp

    odds = pd.read_excel(ODDS)
    odds["team_key_join"] = odds["Selecao"].map(chave_selecao)
    comp = comp.merge(
        odds[["team_key_join", "prob_implicita_media_normalizada"]],
        on="team_key_join", how="left", validate="one_to_one",
    ).rename(columns={"prob_implicita_media_normalizada": "Oddschecker"})

    if comp["Oddschecker"].notna().sum() == 4:
        print("Oddschecker lista somente os quatro semifinalistas; eliminadas permanecem sem cotação.")
    else:
        piso = comp["Oddschecker"].min()
        if pd.notna(piso):
            faltam = comp[comp["Oddschecker"].isna()]["Selecao"].tolist()
            print(f"Times sem Oddschecker (imputados no piso {piso:.6f}): {faltam}")
            comp["Oddschecker"] = comp["Oddschecker"].fillna(piso)
            comp["Oddschecker"] = comp["Oddschecker"] / comp["Oddschecker"].sum()
        else:
            print("Oddschecker sem probabilidades validas. Media_3_fontes seguira sem essa fonte.")
            comp["Oddschecker"] = pd.NA
    return comp


def formatar(caminho):
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Comparativo"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)
    for col in range(2, 7):
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = "0.00%"
            ws[f"{letter}{row}"].alignment = Alignment(horizontal="right")
    ws.conditional_formatting.add(
        f"B2:F{ws.max_row}",
        ColorScaleRule(start_type="min", start_color="EFF6FF", mid_type="percentile",
                       mid_value=50, mid_color="93C5FD", end_type="max", end_color="1D4ED8"),
    )
    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 16
    wb.save(caminho)


def main():
    mercados = pd.read_excel(MERCADOS)
    mercados["team_key_join"] = mercados["Selecao"].map(chave_selecao)
    comp = mercados[["Selecao_PT", "team_key_join", "prob_kalshi_normalizada", "prob_polymarket_normalizada"]].rename(
        columns={"Selecao_PT": "Selecao", "prob_kalshi_normalizada": "Kalshi",
                 "prob_polymarket_normalizada": "Polymarket"}
    )
    comp = carregar_oddschecker(comp)
    fontes = ["Kalshi", "Polymarket", "Oddschecker"]
    comp["Media_3_fontes"] = comp[fontes].mean(axis=1, skipna=True)
    comp["Media_3_fontes"] = comp["Media_3_fontes"] / comp["Media_3_fontes"].sum()

    nao_encontradas = [selecao for selecao in SEM_CHANCE if selecao not in set(comp["Selecao"])]
    if nao_encontradas:
        raise RuntimeError(f"Selecoes a zerar nao encontradas: {nao_encontradas}")
    ajustada = comp["Media_3_fontes"].where(~comp["Selecao"].isin(SEM_CHANCE), 0.0)
    comp["Media_3_ajustada"] = ajustada / ajustada.sum()
    positivas = comp.loc[comp["Media_3_ajustada"] > 0, "Selecao"].tolist()
    print(f"Zeradas ({len(SEM_CHANCE)} eliminados): {SEM_CHANCE}")
    print(f"Positivas apos ajuste ({len(positivas)}): {positivas}")

    cols = ["Selecao", "Kalshi", "Polymarket", "Oddschecker", "Media_3_fontes", "Media_3_ajustada"]
    df = comp[cols].sort_values("Media_3_fontes", ascending=False).reset_index(drop=True)
    df.to_excel(SAIDA, index=False)
    formatar(SAIDA)
    print(f"Salvo: {SAIDA}")
    print("Somas:", {c: round(float(df[c].sum(skipna=True)), 6) for c in cols[1:]})
    print(df.head(48).to_string(index=False))


if __name__ == "__main__":
    main()
