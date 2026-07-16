# -*- coding: utf-8 -*-
"""Gera a simulação oficial da fase 08, início das finais (16/07/2026)."""
from openpyxl import load_workbook

import gerar_sim_oficial_1M_fase07 as fase_anterior


ETAPA = "Início das Finais"
DATA_STR = "16.07.2026"
N_SIMS = 1_000_000
OUT_XLSX = fase_anterior.BASE_DIR / "resultados" / f"Simulação Oficial {ETAPA} {DATA_STR}.xlsx"


def atualizar_metadados():
    wb = load_workbook(OUT_XLSX)
    ws = wb["Parâmetros"]
    valores = {
        "Etapa": ETAPA,
        "Modo do vetor de força": "Ajuste unidimensional: Espanha fixa e Argentina calibrada ao mercado da final",
        "Número de Copas": N_SIMS,
        "Jogos oficiais travados": 102,
    }
    for row in ws.iter_rows(min_row=2, max_col=2):
        parametro = row[0].value
        if parametro in valores:
            row[1].value = valores[parametro]
    wb.save(OUT_XLSX)


def main():
    fase_anterior.ETAPA = ETAPA
    fase_anterior.DATA_STR = DATA_STR
    fase_anterior.N_SIMS = N_SIMS
    fase_anterior.OUT_XLSX = OUT_XLSX
    fase_anterior.main()
    atualizar_metadados()
    print(f"Metadados da fase 08 atualizados: {OUT_XLSX}")


if __name__ == "__main__":
    main()
