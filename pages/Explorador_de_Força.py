from __future__ import annotations

import os
import re
import sys
from io import BytesIO
import json
from datetime import datetime as _dt

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, Alignment, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.helpers import inject_custom_css
from utils.simulador_oficial import simulate_one_cup_oficial, PoissonMatchSimulator
from utils.simulador_analitico import run_detailed_simulation
from utils.forca_core import (
    BASE_DIR,
    DATA_DIR,
    build_combined,
    compute_match_probabilities,
    load_force_dataframe,
    render_param_sidebar,
    team_with_flag,
    TEAM_FLAG_EMOJI,
)


SIM_STAGE_COLUMNS = ["pos_1", "pos_2", "pos_3", "pos_4", "Top32", "Oitavas", "Quartas", "Semifinal", "Final", "Campeao"]


# ----- Helpers de exibição: bandeiras nas tabelas e colunas de probabilidade -----

# Colunas que contêm o nome de UMA seleção (recebem a bandeira como prefixo).
_TEAM_COLS_SINGLE = {
    "Seleção", "Selecao", "Adversario", "Eliminador",
    "Finalista A", "Finalista B", "Bottom 16",
}
# Colunas com vários times num texto único: nome da coluna -> separador.
_TEAM_COLS_MULTI = {"Final": " x "}


def _flag_single(name: object) -> object:
    return team_with_flag(name) if isinstance(name, str) else name


def _flag_multi(text: object, sep: str) -> object:
    if not isinstance(text, str):
        return text
    return sep.join(team_with_flag(part.strip()) for part in text.split(sep))


def add_team_flags(df: pd.DataFrame) -> pd.DataFrame:
    """Devolve uma cópia do DataFrame com o emoji da bandeira antes dos nomes das seleções."""
    out = df.copy()
    for col in out.columns:
        if col in _TEAM_COLS_SINGLE:
            out[col] = out[col].map(_flag_single)
        elif col in _TEAM_COLS_MULTI:
            sep = _TEAM_COLS_MULTI[col]
            out[col] = out[col].map(lambda value, _sep=sep: _flag_multi(value, _sep))
    return out


def pct_col(label: str | None = None, width: str | None = "medium") -> "st.column_config.ProgressColumn":
    """Coluna de probabilidade em % com 1 casa decimal e barra na célula.

    Os valores devem estar em escala 0–100 (use ``prep_pct`` para multiplicar por 100).
    Em tabelas ``width="stretch"`` passe ``width=None`` para não fixar a largura da coluna
    (o stretch distribui o espaço); em tabelas ``content`` o padrão ``"medium"`` mantém a barra legível.
    """
    return st.column_config.ProgressColumn(
        label, format="%.1f%%", width=width, min_value=0, max_value=100
    )


def prep_pct(df: pd.DataFrame, prob_cols: list[str]) -> pd.DataFrame:
    """Cópia do DataFrame com bandeiras nos nomes e colunas de probabilidade em escala 0–100."""
    out = add_team_flags(df)
    for col in prob_cols:
        if col in out.columns:
            out[col] = out[col] * 100
    return out


def final_pair_label(text: object) -> object:
    """Formata o confronto da final: bandeira antes do 1º time e depois do 2º.

    Ex.: "França x Espanha" -> "🇫🇷 França  x  Espanha 🇪🇸" (visual de mando, como num placar).
    """
    if not isinstance(text, str) or " x " not in text:
        return text
    team_a, team_b = (part.strip() for part in text.split(" x ", 1))
    left = team_with_flag(team_a)
    right = f"{team_b} {TEAM_FLAG_EMOJI.get(team_b, '')}".strip()
    return f"{left}  x  {right}"


def show_table(df: pd.DataFrame, height: int = 360, use_container_width: bool = False) -> None:
    """Exibe uma tabela detalhada da simulação.

    Remove colunas auxiliares de base amostral ("Base..."), padroniza a coluna única de
    probabilidade para "Probabilidade" (o contexto fica no subtítulo acima), adiciona as
    bandeiras e formata as probabilidades em % com barra na célula.
    """
    df = df.copy()
    
    # Padronizar a coluna de Seleção para exibição correta
    df = df.rename(columns={"Selecao": "Seleção"})
    
    df = df.drop(
        columns=[col for col in df.columns if str(col).startswith("Base")],
        errors="ignore",
    )
    
    # Identificar se a tabela lista e ordena seleções/adversários/eliminadores
    has_selection_col = any(col in ["Seleção", "Adversario", "Eliminador"] for col in df.columns)
    is_scenario_table = any(col in ["Condicao", "Fase"] for col in df.columns)
    
    if has_selection_col and not is_scenario_table and "Rank" not in df.columns:
        df.insert(0, "Rank", range(1, len(df) + 1))
        
    prob_cols = [col for col in df.columns if str(col).startswith("Prob")]
    
    # Rótulos amigáveis para as colunas de probabilidade específicas
    friendly_labels = {
        "Prob titulo se 1o grupo": "Se 1º no Grupo",
        "Prob titulo se 2o grupo": "Se 2º no Grupo",
        "Prob titulo se avancou em 3o": "Se avançou em 3º",
    }
    
    configs = {}
    for col in prob_cols:
        label = friendly_labels.get(col, None)
        configs[col] = pct_col(label=label)
        
    if len(prob_cols) == 1 and prob_cols[0] not in friendly_labels:
        df = df.rename(columns={prob_cols[0]: "Probabilidade"})
        configs = {"Probabilidade": pct_col()}
        prob_cols = ["Probabilidade"]
        
    if "Rank" in df.columns:
        configs["Rank"] = st.column_config.NumberColumn("Rank", width=50, format="%d")
        
    st.dataframe(
        prep_pct(df, prob_cols),
        use_container_width=use_container_width,
        height=height,
        hide_index=True,
        column_config=configs,
    )


def build_match_cache(
    dataframe: pd.DataFrame,
    media_gols: float,
    usar_dixon_coles: bool,
    rho_dixon_coles: float,
) -> dict[tuple[str, str], dict[str, float]]:
    cache: dict[tuple[str, str], dict[str, float]] = {}
    rows = dataframe.set_index("team_key")[["forca_com_offset"]]
    team_keys = list(rows.index)
    for team_a in team_keys:
        force_a = float(rows.loc[team_a, "forca_com_offset"])
        for team_b in team_keys:
            if team_a == team_b:
                continue
            force_b = float(rows.loc[team_b, "forca_com_offset"])
            cache[(team_a, team_b)] = compute_match_probabilities(
                force_a=force_a,
                force_b=force_b,
                media_gols=media_gols,
                max_goals=10,
                usar_dixon_coles=usar_dixon_coles,
                rho_dixon_coles=rho_dixon_coles,
            )
    return cache


def build_simulation_table(
    dataframe: pd.DataFrame,
    accumulated: dict[str, dict[str, int]],
    n_sims: int,
) -> pd.DataFrame:
    result = dataframe[
        [
            "Seleção",
            "team_key",
            "forca_com_offset",
        ]
    ].copy()

    for stage in SIM_STAGE_COLUMNS:
        result[f"{stage}_pct"] = result["team_key"].map(lambda key: accumulated[key][stage] / n_sims)

    # Ordenar pelos favoritos (Campeão, Final, etc.)
    result = result.sort_values(
        by=["Campeao_pct", "Final_pct", "Semifinal_pct", "forca_com_offset"],
        ascending=False,
    ).reset_index(drop=True)

    result.index = result.index + 1
    result.insert(0, "Rank Sim", result.index)

    return result


# Fases eliminatórias usadas nas tabelas agregadas por categoria e nº de vagas em cada fase.
# Somando a probabilidade por seleção e dividindo pelas vagas, cada coluna vira uma
# "participação esperada" da categoria na fase (todas as colunas somam 100% entre categorias).
CATEGORY_STAGE_SLOTS = {
    "Top32": 32,
    "Oitavas": 16,
    "Quartas": 8,
    "Semifinal": 4,
    "Final": 2,
    "Campeao": 1,
}

CATEGORY_STAGE_LABELS = {
    "Top32": "Top 32",
    "Oitavas": "Oitavas",
    "Quartas": "Quartas",
    "Semifinal": "Semi",
    "Final": "Final",
    "Campeao": "Campeão",
}


def _aggregate_category(
    merged: pd.DataFrame,
    group_col: str,
    sort_alpha: bool = False,
    order_key=None,
) -> pd.DataFrame:
    """Agrega as probabilidades por fase somando entre as seleções de cada categoria.

    Cada coluna de fase é dividida pelo nº de vagas da fase, virando a fração esperada
    daquela fase ocupada pela categoria (a coluna ``Campeão`` é literalmente a
    probabilidade de o campeão sair da categoria).
    """
    rows = []
    for category, sub in merged.groupby(group_col):
        row = {
            "Categoria": category,
            "Nº Seleções": int(len(sub)),
            "Força Média": float(sub["forca_com_offset"].mean()),
        }
        for stage, slots in CATEGORY_STAGE_SLOTS.items():
            row[CATEGORY_STAGE_LABELS[stage]] = float(sub[f"{stage}_pct"].sum()) / slots
        rows.append(row)

    result = pd.DataFrame(rows)
    if "Campeão" in result.columns and "Nº Seleções" in result.columns:
        result["Média Campeão"] = result["Campeão"] / result["Nº Seleções"]

    if order_key is not None:
        result["_ord"] = result["Categoria"].map(order_key)
        result = result.sort_values(by="_ord").drop(columns="_ord")
    elif sort_alpha:
        result = result.sort_values(by="Categoria")
    else:
        result = result.sort_values(by="Campeão", ascending=False)
    return result.reset_index(drop=True)


def _contar_titulos_mundiais(melhor: object) -> int:
    """Conta quantos títulos mundiais a seleção tem a partir de 'Campeão (anos...)'."""
    texto = str(melhor or "").strip()
    if not texto.startswith("Campe"):
        return 0
    match = re.search(r"\(([^)]*)\)", texto)
    if not match:
        return 1
    anos = [ano for ano in re.split(r"[,/;]", match.group(1)) if ano.strip()]
    return len(anos) if anos else 1


def _ordem_titulos(categoria: str) -> tuple[int, int]:
    """Ordem complementar: estreantes, nunca campeãs e depois 1, 2, 3... títulos."""
    if categoria.startswith("Estre"):
        return (0, 0)
    if categoria.startswith("Nunca"):
        return (1, 0)
    match = re.match(r"(\d+)", categoria)
    return (2, int(match.group(1)) if match else 99)


# Colunas da distribuição de eliminação, na ordem das fases (somam 100% por seleção).
ELIMINATION_COLUMNS = [
    "Fase de Grupos",
    "16-avos",
    "Oitavas",
    "Quartas",
    "Semifinal",
    "Vice (Final)",
    "Campeã",
]


def build_elimination_table(sim_table: pd.DataFrame) -> pd.DataFrame:
    """Distribuição da fase em que cada seleção é eliminada.

    Derivada das probabilidades acumuladas ("chegou pelo menos à fase X"): a chance de
    sair exatamente numa fase é a diferença entre chegar nela e chegar na seguinte.
    Cada linha soma ~100%.
    """
    df = sim_table
    out = pd.DataFrame({"Seleção": df["Seleção"].values})
    out["Fase de Grupos"] = 1.0 - df["Top32_pct"].values
    out["16-avos"] = df["Top32_pct"].values - df["Oitavas_pct"].values
    out["Oitavas"] = df["Oitavas_pct"].values - df["Quartas_pct"].values
    out["Quartas"] = df["Quartas_pct"].values - df["Semifinal_pct"].values
    out["Semifinal"] = df["Semifinal_pct"].values - df["Final_pct"].values
    out["Vice (Final)"] = df["Final_pct"].values - df["Campeao_pct"].values
    out["Campeã"] = df["Campeao_pct"].values

    # Arredondamentos de ponto flutuante podem gerar diferenças levemente negativas.
    for col in ELIMINATION_COLUMNS:
        out[col] = out[col].clip(lower=0.0)

    out = out.sort_values(
        by=["Campeã", "Vice (Final)", "Semifinal", "Quartas"], ascending=False
    ).reset_index(drop=True)
    out.index = out.index + 1
    out.insert(0, "Rank", out.index)
    return out


GROUP_STAGE_PROB_COLUMNS = [
    "1º",
    "2º",
    "3º",
    "4º",
    "Avança como 3º",
    "Classifica (Top 32)",
    "Cai na fase de grupos",
]


def build_group_stage_table(
    sim_table: pd.DataFrame, meta_df: pd.DataFrame
) -> pd.DataFrame:
    """Detalhamento da fase de grupos por seleção, organizado por grupo.

    Tudo derivado das probabilidades já acumuladas: ``Avança como 3º`` é a parcela do
    Top 32 que sobra depois de 1º e 2º (os 8 melhores terceiros), e ``Cai na fase de
    grupos`` é o complemento do Top 32.
    """
    merged = sim_table.merge(meta_df[["team_key", "Grupo"]], on="team_key", how="left")

    out = pd.DataFrame({"Grupo": merged["Grupo"].values, "Seleção": merged["Seleção"].values})
    out["1º"] = merged["pos_1_pct"].values
    out["2º"] = merged["pos_2_pct"].values
    out["3º"] = merged["pos_3_pct"].values
    out["4º"] = merged["pos_4_pct"].values
    out["Avança como 3º"] = (
        merged["Top32_pct"].values - merged["pos_1_pct"].values - merged["pos_2_pct"].values
    )
    out["Classifica (Top 32)"] = merged["Top32_pct"].values
    out["Cai na fase de grupos"] = 1.0 - merged["Top32_pct"].values

    for col in GROUP_STAGE_PROB_COLUMNS:
        out[col] = out[col].clip(lower=0.0)

    out = out.sort_values(
        by=["Grupo", "Classifica (Top 32)"], ascending=[True, False]
    ).reset_index(drop=True)
    return out


def build_category_tables(
    sim_table: pd.DataFrame, meta_df: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    """Constrói as tabelas de probabilidade agregadas por grupo, confederação,
    estreia em Copas e tradição de título a partir da tabela de simulação."""
    meta_cols = [
        "team_key",
        "Grupo",
        "Confederação",
        "Participações_Copa_Mundo",
        "Melhor_Resultado_Copa_Mundo",
    ]
    merged = sim_table.merge(meta_df[meta_cols], on="team_key", how="left")

    merged["cat_grupo"] = "Grupo " + merged["Grupo"].astype(str)
    merged["cat_confed"] = merged["Confederação"].fillna("Sem confederação")

    apps = pd.to_numeric(merged["Participações_Copa_Mundo"], errors="coerce").fillna(-1)
    titulos = merged["Melhor_Resultado_Copa_Mundo"].apply(_contar_titulos_mundiais)

    def _cat_titulos(n_apps: float, n_titulos: int) -> str:
        if n_apps == 0:
            return "Estreantes (1ª Copa)"
        if n_titulos == 0:
            return "Nunca campeãs"
        if n_titulos in [1, 2]:
            return "1 ou 2 títulos"
        if n_titulos >= 3:
            return "3+ títulos"
        return "Nunca campeãs"

    merged["cat_titulos"] = [_cat_titulos(a, t) for a, t in zip(apps, titulos)]

    return {
        "Por grupo": _aggregate_category(merged, "cat_grupo", sort_alpha=True),
        "Por confederação": _aggregate_category(merged, "cat_confed"),
        "Por títulos em Copas": _aggregate_category(merged, "cat_titulos", order_key=_ordem_titulos),
    }


_MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}


def _parse_date_pt(raw: str) -> str:
    """'Quinta-feira, 11 de junho de 2026' -> '11/06/2026' (sortable in Excel)."""
    m = re.search(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", raw)
    if m:
        day = int(m.group(1))
        month = _MESES_PT.get(m.group(2).lower(), 0)
        year = int(m.group(3))
        if month:
            return _dt(year, month, day).strftime("%d/%m/%Y")
    return raw


def _extract_local(raw: str) -> str:
    """'Cidade do México, no México — 13h00 no horário local' -> 'Cidade do México, México'."""
    # Remove tudo a partir de ' —' ou ' –' (travessão)
    local = re.split(r"\s*[—–]\s*", raw)[0].strip()
    # Limpar ', no/nos/na/nas PAIS' -> ', PAIS'
    local = re.sub(r",\s*n[oa]s?\s+", ", ", local)
    return local


def _extract_br_time(raw: str) -> str:
    """'(16h00 em Brasília / 18h00 em Praia / ...)' -> '16h00'.
    Also handles '(1h00 de 14 de junho em Brasília ...) -> '01h00'."""
    m = re.search(r"\((\d{1,2}h\d{2})", raw)
    if m:
        t = m.group(1)
        # Normalizar hora de 1 dígito -> 2 dígitos (1h00 -> 01h00)
        parts = t.split("h")
        return f"{int(parts[0]):02d}h{parts[1]}"
    return raw


def generate_group_predictions(
    dataframe: pd.DataFrame, media_gols: float, usar_dixon_coles: bool, rho_dixon_coles: float
) -> pd.DataFrame:
    try:
        with open(DATA_DIR / "calendario_copa_2026.json", "r", encoding="utf-8") as f:
            schedule = json.load(f)
    except FileNotFoundError:
        return pd.DataFrame()

    name_map = {
        "República da Coreia": "Coreia do Sul",
        "República Democrática do Congo": "RD do Congo",
        "República Tcheca": "Tcheca"
    }

    rows = []
    for match in schedule:
        team_a_str = name_map.get(match["team_a"], match["team_a"])
        team_b_str = name_map.get(match["team_b"], match["team_b"])

        # Obter forca se times existirem
        if team_a_str in dataframe["Seleção"].values and team_b_str in dataframe["Seleção"].values:
            forca_a = float(dataframe.loc[dataframe["Seleção"] == team_a_str, "forca_com_offset"].iloc[0])
            forca_b = float(dataframe.loc[dataframe["Seleção"] == team_b_str, "forca_com_offset"].iloc[0])

            # A função compute_match_probabilities já usa Poisson, média e rho do slider
            probs = compute_match_probabilities(
                force_a=forca_a, force_b=forca_b, media_gols=media_gols,
                usar_dixon_coles=usar_dixon_coles, rho_dixon_coles=rho_dixon_coles
            )

            wa, d, wb = probs["win_a"], probs["draw"], probs["win_b"]

            rows.append({
                "Grupo": f"Grupo {match['group']}",
                "Data": _parse_date_pt(match["date"]),
                "Local": _extract_local(match["location_time"]),
                "Horário Brasília": _extract_br_time(match["br_time"]),
                "Seleção A": team_a_str,
                "Vitória A": wa,
                "Empate": d,
                "Vitória B": wb,
                "Seleção B": team_b_str,
                "Força A": forca_a
            })

    df_res = pd.DataFrame(rows)
    if not df_res.empty:
        df_res = df_res.sort_values(by=["Grupo", "Força A"], ascending=[True, False]).reset_index(drop=True)
        df_res = df_res.drop(columns=["Força A"])
    return df_res


def simulation_excel_bytes(
    simulation_df: pd.DataFrame,
    info_df: pd.DataFrame,
    matches_df: pd.DataFrame = None,
    extra_sheets: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        simulation_df.to_excel(writer, sheet_name="Simulações", index=False)
        info_df.to_excel(writer, sheet_name="Parâmetros", index=False)

        if matches_df is not None and not matches_df.empty:
            # Formatar percentuais para exibicao sem converter pra string pra não estragar
            # Mas pandas to_excel precisa ou da string ou que apliquemos style.
            export_matches = matches_df.copy()
            for col in ["Vitória A", "Empate", "Vitória B"]:
                export_matches[col] = export_matches[col].apply(lambda x: f"{x:.1%}")

            export_matches.to_excel(writer, sheet_name="Previsão Jogos", index=False)
            workbook = writer.book
            worksheet = writer.sheets["Previsão Jogos"]

            header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal="center", vertical="center")

            # Novo esquema de cores
            base_fill = PatternFill(start_color="EBF4FA", end_color="EBF4FA", fill_type="solid") # Azul bem clarinho para toda a planilha
            win_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid") # Azul um pouco mais forte para vitórias
            draw_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid") # Cinza claro para empate

            for col in range(1, 10):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            worksheet.column_dimensions['A'].width = 10   # Grupo
            worksheet.column_dimensions['B'].width = 14   # Data (dd/mm/yyyy)
            worksheet.column_dimensions['C'].width = 35   # Local (cidade, país)
            worksheet.column_dimensions['D'].width = 16   # Horário Brasília (HHhMM)
            worksheet.column_dimensions['E'].width = 20   # Seleção A
            worksheet.column_dimensions['F'].width = 12   # Vitória A
            worksheet.column_dimensions['G'].width = 12   # Empate
            worksheet.column_dimensions['H'].width = 12   # Vitória B
            worksheet.column_dimensions['I'].width = 20   # Seleção B

            for row in range(2, len(export_matches) + 2):
                for col in range(1, 10):
                    # Preenchimento base (azul claro)
                    worksheet.cell(row=row, column=col).fill = base_fill

                    if col in [1, 2, 3, 4, 6, 7, 8]:
                        worksheet.cell(row=row, column=col).alignment = center_align

                worksheet.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
                worksheet.cell(row=row, column=5).font = Font(bold=True)
                worksheet.cell(row=row, column=9).alignment = Alignment(horizontal="left", vertical="center")
                worksheet.cell(row=row, column=9).font = Font(bold=True)

                worksheet.cell(row=row, column=6).fill = win_fill
                worksheet.cell(row=row, column=7).fill = draw_fill
                worksheet.cell(row=row, column=8).fill = win_fill

        if extra_sheets:
            for sheet_name, sheet_df in extra_sheets.items():
                safe_name = sheet_name[:31]
                sheet_df.to_excel(writer, sheet_name=safe_name, index=False)

    buffer.seek(0)
    return buffer.getvalue()


def run_complete_simulation_progressive(
    dataframe: pd.DataFrame,
    media_gols: float,
    n_sims: int,
    usar_dixon_coles: bool,
    rho_dixon_coles: float,
    tipo_chaveamento: str = "Sorteio Oficial",
    chunk_size: int = 10000,
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    strengths = dict(zip(dataframe["team_key"], dataframe["forca_com_offset"]))
    match_cache = build_match_cache(
        dataframe=dataframe,
        media_gols=media_gols,
        usar_dixon_coles=usar_dixon_coles,
        rho_dixon_coles=rho_dixon_coles,
    )
    match_simulator = PoissonMatchSimulator(match_cache=match_cache, strengths=strengths)
    status_placeholder = st.empty()

    def update_progress(completed: int, total: int) -> None:
        status_placeholder.markdown(f"**Progresso detalhado:** {completed:,} / {total:,} copas")

    with st.spinner("Processando simulação completa...", show_time=True):
        detailed_result = run_detailed_simulation(
            dataframe=dataframe,
            n_sims=n_sims,
            match_simulator=match_simulator,
            strengths=strengths,
            tipo_chaveamento=tipo_chaveamento,
            chunk_size=chunk_size,
            progress_callback=update_progress,
        )

    status_placeholder.success(f"Simulação completa concluída: {n_sims:,} copas!")
    sim_table = build_simulation_table(
        dataframe=dataframe,
        accumulated=detailed_result["accumulated"],
        n_sims=n_sims,
    )
    return sim_table, detailed_result["tables"]


def load_saved_simulation(file_path: str | os.PathLike, combined_df: pd.DataFrame) -> bool:
    import unicodedata

    def normalize_str(s: str) -> str:
        return "".join(
            c for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        ).lower()

    def find_sheet(sheet_names, target_normalized):
        for name in sheet_names:
            if normalize_str(name) == target_normalized:
                return name
        for name in sheet_names:
            if target_normalized in normalize_str(name):
                return name
        return None

    try:
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names

        sim_sheet = find_sheet(sheet_names, "simulacoes")
        if not sim_sheet:
            st.error("Planilha 'Simulações' não encontrada no arquivo Excel.")
            return False

        sim_display = pd.read_excel(xls, sim_sheet)

        # Normalizar colunas de sim_display
        col_mapping = {
            "selecao": "Seleção",
            "1 grupo": "1º Grupo",
            "2 grupo": "2º Grupo",
            "3 grupo": "3º Grupo",
            "4 grupo": "4º Grupo",
            "top 32": "Top 32",
            "oitavas": "Oitavas",
            "quartas": "Quartas",
            "semi": "Semi",
            "final": "Final",
            "campeao": "Campeão",
        }
        new_cols = {}
        for col in sim_display.columns:
            norm_col = normalize_str(str(col))
            if norm_col in col_mapping:
                new_cols[col] = col_mapping[norm_col]
        sim_display = sim_display.rename(columns=new_cols)

        # Garantir a presença de Seleção
        if "Seleção" not in sim_display.columns:
            for col in sim_display.columns:
                if "selec" in normalize_str(str(col)):
                    sim_display = sim_display.rename(columns={col: "Seleção"})
                    break

        # Reconstruir sim_table para gerar as tabelas agregadas caso necessário
        sim_table = pd.DataFrame()
        sim_table["Seleção"] = sim_display["Seleção"]

        team_key_map = dict(zip(combined_df["Seleção"], combined_df["team_key"]))
        sim_table["team_key"] = sim_table["Seleção"].map(lambda x: team_key_map.get(x, x))

        forca_map = dict(zip(combined_df["team_key"], combined_df["forca_com_offset"]))
        sim_table["forca_com_offset"] = sim_table["team_key"].map(lambda x: forca_map.get(x, 0.0))

        mapping_rev = {
            "pos_1_pct": "1º Grupo",
            "pos_2_pct": "2º Grupo",
            "pos_3_pct": "3º Grupo",
            "pos_4_pct": "4º Grupo",
            "Top32_pct": "Top 32",
            "Oitavas_pct": "Oitavas",
            "Quartas_pct": "Quartas",
            "Semifinal_pct": "Semi",
            "Final_pct": "Final",
            "Campeao_pct": "Campeão",
        }
        for dest, src in mapping_rev.items():
            if src in sim_display.columns:
                val = sim_display[src]
                if val.max() > 1.0:
                    sim_table[dest] = val / 100.0
                else:
                    sim_table[dest] = val
            else:
                sim_table[dest] = 0.0

        # Gerar tabelas de categoria e de eliminação a partir da sim_table
        elimination_table = build_elimination_table(sim_table)
        group_stage_table = build_group_stage_table(sim_table, combined_df)
        category_tables = build_category_tables(sim_table, combined_df)

        # Reconstruir detailed_tables
        detailed_tables = {}
        sheet_mapping = {
            "Finais": "finais",
            "Brasil 1o Top32": "brasil_1o_grupo_top32",
            "Brasil 2o Top32": "brasil_2o_grupo_top32",
            "Brasil 3o Top32": "brasil_3o_grupo_top32",
            "Brasil Adv 16avos": "brasil_adversarios_16avos",
            "Brasil Adv Oitavas": "brasil_adversarios_oitavas",
            "Brasil Adv Quartas": "brasil_adversarios_quartas",
            "Brasil Adv Semi": "brasil_adversarios_semifinal",
            "Brasil Adv Final": "brasil_adversarios_final",
            "Eliminadores Brasil": "eliminadores_brasil",
            "Carrascos Brasil": "eliminadores_brasil_agrupado",
            "Titulo Cond Brasil": "titulo_condicional_brasil",
            "Impacto Pos Grupo": "impacto_posicao_grupo",
            "Bottom16 Surpresa": "bottom16_surpresa",
            "Bottom16 Lista": "bottom16_lista",
            "MiniZebra Surpresa": "minizebra_surpresa",
            "MiniZebra Lista": "minizebra_lista",
        }

        has_any_detailed = False
        for sheet_name, dict_key in sheet_mapping.items():
            norm_target = normalize_str(sheet_name)
            found_sheet = find_sheet(sheet_names, norm_target)
            if found_sheet:
                df_sheet = pd.read_excel(xls, found_sheet)
                df_sheet = df_sheet.rename(columns={
                    "Selecao": "Seleção",
                    "Condicao": "Condição",
                })
                detailed_tables[dict_key] = df_sheet
                has_any_detailed = True
            else:
                detailed_tables[dict_key] = pd.DataFrame()

        with open(file_path, "rb") as f:
            excel_bytes = f.read()

        st.session_state["explorador_sim_display"] = sim_display
        st.session_state["explorador_detailed_tables"] = detailed_tables if has_any_detailed else None
        st.session_state["explorador_category_tables"] = category_tables
        st.session_state["explorador_elimination_table"] = elimination_table
        st.session_state["explorador_group_stage_table"] = group_stage_table
        st.session_state["explorador_sim_excel_bytes"] = excel_bytes
        st.session_state["explorador_loaded_filename"] = os.path.basename(file_path)

        return True
    except Exception as e:
        st.error(f"Erro ao carregar simulação: {e}")
        return False


inject_custom_css()

st.markdown("## Simulação Copa do Mundo 2026")
st.markdown(
    """
<p style="font-size: 1rem; margin-bottom: 1.5rem;">
Simulação completa da Copa do Mundo de 2026 a partir do indicador de força e dos
parâmetros do modelo definidos na barra lateral. Ajuste o número de copas e o tipo
de chaveamento abaixo e rode a simulação.
</p>
""",
    unsafe_allow_html=True,
)

params = render_param_sidebar()
base_df = load_force_dataframe()
combined_df, weight_sum = build_combined(base_df, params)

if weight_sum <= 0:
    st.warning("A soma dos pesos está zerada. Ajuste ao menos um peso na barra lateral para construir a força resultante.")

media_gols = params.media_gols
usar_dixon_coles = params.usar_dixon_coles
rho_dixon_coles = params.rho_dixon_coles
tipo_simulacao = "Completa"

if "explorador_sim_excel_bytes" not in st.session_state:
    st.session_state["explorador_sim_excel_bytes"] = None

st.markdown("### Parâmetros da Simulação")


def _fmt_copas(v: int) -> str:
    """Formata o nº de copas para o rótulo do slider (ex.: 500000 -> '500 mil', 5000000 -> '5 milhões')."""
    if v >= 1_000_000:
        mi = v / 1_000_000
        num = f"{mi:.0f}" if mi == int(mi) else f"{mi:.1f}"
        unidade = "milhão" if mi == 1 else "milhões"
        return f"{num} {unidade}"
    return f"{v // 1000} mil"


SIM_COPAS_OPCOES = (
    list(range(10000, 100001, 10000))        # 10 mil → 100 mil (de 10 mil em 10 mil)
    + list(range(200000, 1000001, 100000))   # 200 mil → 1 mi (de 100 mil em 100 mil)
    + list(range(2000000, 10000001, 1000000))  # 2 mi → 10 mi (de 1 mi em 1 mi)
)

tab_rodar, tab_carregar = st.tabs(["⚙️ Ajustar e Rodar Nova Simulação", "📂 Carregar Simulação Oficial ou Salva"])

with tab_rodar:
    col_a, col_b = st.columns([1, 1])
    with col_a:
        n_sims = st.select_slider(
            "Nº de Copas",
            options=SIM_COPAS_OPCOES,
            value=10000,
            format_func=_fmt_copas,
            key="sim_n_sims_preset",
        )
    with col_b:
        tipo_chaveamento = st.pills(
            "Chaveamento", ["Sorteio Oficial", "Sorteio Aleatório"], selection_mode="single", default="Sorteio Oficial", key="sim_tipo_chaveamento"
        )
        if tipo_chaveamento is None:
            tipo_chaveamento = "Sorteio Oficial"

    # ETA estimado considerando ~1000 simulações por segundo
    eta_min = int(n_sims) / 1000 / 60
    eta_label = f"{eta_min:.1f}".replace(".", ",")
    run_simulation = st.button(
        f"🚀 Rodar simulação (ETA: {eta_label}min)", type="primary", use_container_width=True
    )

with tab_carregar:
    resultados_dir = BASE_DIR / "resultados"
    os.makedirs(resultados_dir, exist_ok=True)
    saved_files = [f for f in os.listdir(resultados_dir) if f.endswith(".xlsx")]

    if not saved_files:
        st.info("Nenhum arquivo `.xlsx` encontrado na pasta `resultados/`.")
        load_simulation = False
    else:
        # Colocar o arquivo oficial ou mais recente primeiro
        saved_files_sorted = sorted(
            saved_files,
            key=lambda x: ("pre-torneio" in x.lower() or "oficial" in x.lower(), x),
            reverse=True
        )
        col_sel, col_btn = st.columns([3, 1])
        with col_sel:
            selected_file = st.selectbox(
                "Selecione uma simulação salva:",
                options=saved_files_sorted,
                help="Carregue uma simulação completa previamente executada para visualizar as abas imediatamente.",
                key="sim_selected_file_to_load"
            )
        with col_btn:
            # Spacer to vertically align button with selectbox input
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            load_simulation = st.button(
                "📥 Carregar",
                type="primary",
                use_container_width=True
            )

        if load_simulation:
            file_path = resultados_dir / selected_file
            if load_saved_simulation(file_path, combined_df):
                st.success(f"Resultados de '{selected_file}' carregados com sucesso!")
                st.rerun()

st.markdown("---")
st.markdown("### Simulação")

if "explorador_sim_display" not in st.session_state:
    st.session_state["explorador_sim_display"] = None
if "explorador_detailed_tables" not in st.session_state:
    st.session_state["explorador_detailed_tables"] = None
if "explorador_category_tables" not in st.session_state:
    st.session_state["explorador_category_tables"] = None
if "explorador_elimination_table" not in st.session_state:
    st.session_state["explorador_elimination_table"] = None
if "explorador_group_stage_table" not in st.session_state:
    st.session_state["explorador_group_stage_table"] = None
if "explorador_loaded_filename" not in st.session_state:
    st.session_state["explorador_loaded_filename"] = None

loaded_filename = st.session_state.get("explorador_loaded_filename")
if loaded_filename:
    st.info(f"ℹ️ **Exibindo resultados da simulação carregada:** `{loaded_filename}`")

if run_simulation:
    if "explorador_loaded_filename" in st.session_state:
        st.session_state["explorador_loaded_filename"] = None
    sim_table, detailed_tables = run_complete_simulation_progressive(
        dataframe=combined_df,
        media_gols=media_gols,
        n_sims=int(n_sims),
        usar_dixon_coles=usar_dixon_coles,
        rho_dixon_coles=rho_dixon_coles,
        tipo_chaveamento=tipo_chaveamento,
        chunk_size=10000,
    )

    sim_display = sim_table[
        [
            "Rank Sim",
            "Seleção",
            "pos_1_pct",
            "pos_2_pct",
            "pos_3_pct",
            "pos_4_pct",
            "Top32_pct",
            "Oitavas_pct",
            "Quartas_pct",
            "Semifinal_pct",
            "Final_pct",
            "Campeao_pct",
        ]
    ].rename(
        columns={
            "Rank Sim": "Rank",
            "pos_1_pct": "1º Grupo",
            "pos_2_pct": "2º Grupo",
            "pos_3_pct": "3º Grupo",
            "pos_4_pct": "4º Grupo",
            "Top32_pct": "Top 32",
            "Oitavas_pct": "Oitavas",
            "Quartas_pct": "Quartas",
            "Semifinal_pct": "Semi",
            "Final_pct": "Final",
            "Campeao_pct": "Campeão",
        }
    )

    info_df = pd.DataFrame(
        [
            {"Parametro": "Etapa", "Valor": "Pré-Torneio"},
            {"Parametro": "Peso FIFA", "Valor": params.weight_fifa},
            {"Parametro": "Peso ELO", "Valor": params.weight_elo},
            {"Parametro": "Peso Momento", "Valor": params.weight_momentum},
            {"Parametro": "Peso Mercado", "Valor": params.weight_market},
            {"Parametro": "Peso Histórico Copas", "Valor": params.weight_history},
            {"Parametro": "Peso Anfitrião", "Valor": params.weight_host},
            {"Parametro": "Offset", "Valor": params.offset},
            {"Parametro": "Elasticidade", "Valor": params.elasticidade},
            {"Parametro": "Média de gols", "Valor": media_gols},
            {"Parametro": "Usar Dixon-Coles", "Valor": usar_dixon_coles},
            {"Parametro": "Rho Dixon-Coles", "Valor": rho_dixon_coles},
            {"Parametro": "Número de Copas", "Valor": int(n_sims)},
            {"Parametro": "Tipo de Simulação", "Valor": tipo_simulacao},
        ]
    )

    matches_df = generate_group_predictions(
        dataframe=combined_df, media_gols=media_gols,
        usar_dixon_coles=usar_dixon_coles, rho_dixon_coles=rho_dixon_coles
    )

    elimination_table = build_elimination_table(sim_table)
    group_stage_table = build_group_stage_table(sim_table, combined_df)
    category_tables = build_category_tables(sim_table, combined_df)

    extra_sheets = None
    if detailed_tables:
        extra_sheets = {
            "Finais": detailed_tables["finais"],
            "Brasil 1o Top32": detailed_tables["brasil_1o_grupo_top32"],
            "Brasil 2o Top32": detailed_tables["brasil_2o_grupo_top32"],
            "Brasil 3o Top32": detailed_tables["brasil_3o_grupo_top32"],
            "Brasil Adv 16avos": detailed_tables["brasil_adversarios_16avos"],
            "Brasil Adv Oitavas": detailed_tables["brasil_adversarios_oitavas"],
            "Brasil Adv Quartas": detailed_tables["brasil_adversarios_quartas"],
            "Brasil Adv Semi": detailed_tables["brasil_adversarios_semifinal"],
            "Brasil Adv Final": detailed_tables["brasil_adversarios_final"],
            "Eliminadores Brasil": detailed_tables["eliminadores_brasil"],
            "Carrascos Brasil": detailed_tables["eliminadores_brasil_agrupado"],
            "Titulo Cond Brasil": detailed_tables["titulo_condicional_brasil"],
            "Impacto Pos Grupo": detailed_tables["impacto_posicao_grupo"],
            "Bottom16 Surpresa": detailed_tables["bottom16_surpresa"],
            "Bottom16 Lista": detailed_tables["bottom16_lista"],
            "MiniZebra Surpresa": detailed_tables["minizebra_surpresa"],
            "MiniZebra Lista": detailed_tables["minizebra_lista"],
        }

    if extra_sheets is None:
        extra_sheets = {}
    extra_sheets["Fase de Grupos Detalhe"] = group_stage_table
    extra_sheets["Fase de Eliminacao"] = elimination_table
    extra_sheets["Cat Por Grupo"] = category_tables["Por grupo"]
    extra_sheets["Cat Por Confederacao"] = category_tables["Por confederação"]
    extra_sheets["Cat Por Titulos"] = category_tables["Por títulos em Copas"]

    st.session_state["explorador_sim_excel_bytes"] = simulation_excel_bytes(
        sim_display,
        info_df,
        matches_df,
        extra_sheets=extra_sheets,
    )

    if int(n_sims) >= 100000:
        import datetime
        os.makedirs(BASE_DIR / "resultados", exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%d-%m-%Y")
        filepath = BASE_DIR / "resultados" / f"simulacao_previsao_esportiva_{tipo_simulacao}_Pre-Torneio_{timestamp}.xlsx"
        with open(filepath, "wb") as f:
            f.write(st.session_state["explorador_sim_excel_bytes"])

    st.session_state["explorador_sim_display"] = sim_display
    st.session_state["explorador_detailed_tables"] = detailed_tables
    st.session_state["explorador_category_tables"] = category_tables
    st.session_state["explorador_elimination_table"] = elimination_table
    st.session_state["explorador_group_stage_table"] = group_stage_table
    st.rerun()

sim_display_state = st.session_state.get("explorador_sim_display")
group_stage_table = st.session_state.get("explorador_group_stage_table")
elimination_table = st.session_state.get("explorador_elimination_table")
detailed_tables = st.session_state.get("explorador_detailed_tables")
category_tables = st.session_state.get("explorador_category_tables")

_HELP_P = '<p style="font-size: 0.95rem; color: #555; margin-bottom: 1rem;">{}</p>'

if sim_display_state is not None:
    # Garantir que a coluna 'Rank' existe no estado recuperado (caso seja de sessão antiga)
    if "Rank" not in sim_display_state.columns:
        if "Rank Sim" in sim_display_state.columns:
            sim_display_state = sim_display_state.rename(columns={"Rank Sim": "Rank"})
        else:
            sim_display_state = sim_display_state.copy()
            sim_display_state.insert(0, "Rank", range(1, len(sim_display_state) + 1))

    # Criar abas para as seções de resultado
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "🏆 Probabilidade por Seleção",
        "😭 Probabilidades de Eliminação",
        "🇧🇷 Caminho do Brasil",
        "🏁 Finais Mais Prováveis",
        "🦓 Zebras (Surpresas)",
        "📁 Probabilidades por Categoria"
    ])

    with tab1:
        # ===================== 1. Probabilidade por seleção =====================
        st.markdown("### 🏆 Probabilidade por seleção")
        st.markdown(
            _HELP_P.format(
                "Ranking pelos favoritos ao título. Para cada seleção, a probabilidade de terminar "
                "em cada posição do grupo (<b>1º</b>–<b>4º</b>) e de alcançar cada fase do mata-mata, "
                "até ser <b>Campeão</b>."
            ),
            unsafe_allow_html=True,
        )
        _main_prob_cols = [
            "1º Grupo", "2º Grupo", "3º Grupo", "4º Grupo", "Top 32",
            "Oitavas", "Quartas", "Semi", "Final", "Campeão",
        ]
        st.dataframe(
            prep_pct(sim_display_state, _main_prob_cols),
            use_container_width=True,
            height=520,
            hide_index=True,
            column_config={
                "Rank": st.column_config.NumberColumn("Rank", width=50, format="%d"),
                "Seleção": st.column_config.TextColumn("Seleção", width="medium"),
                **{col: pct_col(width="small") for col in _main_prob_cols}
            },
        )

        if detailed_tables:
            st.markdown("#### Título conforme a posição no grupo")
            st.markdown(
                _HELP_P.format(
                    "Para cada seleção, a probabilidade de ser campeã conforme tenha terminado em "
                    "<b>1º</b>, <b>2º</b> ou avançado em <b>3º</b> do grupo."
                ),
                unsafe_allow_html=True,
            )
            show_table(detailed_tables["impacto_posicao_grupo"], height=520)

    with tab2:
        # ===================== 2. Probabilidades de Eliminação =====================
        st.markdown("### 😭 Probabilidades de eliminação")
        st.markdown(
            _HELP_P.format(
                "Distribuição da probabilidade de eliminação de cada seleção em cada fase do mata-mata."
            ),
            unsafe_allow_html=True,
        )

        if elimination_table is not None:
            # Garantir que a coluna 'Rank' existe no estado recuperado da tabela de eliminação
            if "Rank" not in elimination_table.columns:
                elimination_table = elimination_table.copy()
                elimination_table.insert(0, "Rank", range(1, len(elimination_table) + 1))

            st.markdown("#### Onde cada seleção é eliminada")
            st.markdown(
                _HELP_P.format(
                    "Em qual fase cada seleção é eliminada — da fase de grupos ao título. Cada linha "
                    "soma 100%: <b>16-avos</b> = chegou ao mata-mata de 32 e perdeu; <b>Vice</b> = "
                    "perdeu a final; <b>Campeã</b> = venceu o torneio."
                ),
                unsafe_allow_html=True,
            )
            st.dataframe(
                prep_pct(elimination_table, ELIMINATION_COLUMNS),
                use_container_width=True,
                height=520,
                hide_index=True,
                column_config={
                    "Rank": st.column_config.NumberColumn("Rank", width=50, format="%d"),
                    "Seleção": st.column_config.TextColumn("Seleção", width="medium"),
                    **{col: pct_col(width=None) for col in ELIMINATION_COLUMNS},
                },
            )

    with tab3:
        # ===================== 3. Caminho do Brasil =====================
        if detailed_tables:
            st.markdown("### 🇧🇷 Caminho do Brasil")

            st.markdown("#### Adversário no 1º mata-mata, por posição no grupo")
            col_brasil_1, col_brasil_2, col_brasil_3 = st.columns(3)
            with col_brasil_1:
                st.caption("Brasil em 1º do grupo")
                show_table(detailed_tables["brasil_1o_grupo_top32"], height=320)
            with col_brasil_2:
                st.caption("Brasil em 2º do grupo")
                show_table(detailed_tables["brasil_2o_grupo_top32"], height=320)
            with col_brasil_3:
                st.caption("Brasil avançou em 3º")
                show_table(detailed_tables["brasil_3o_grupo_top32"], height=320)

            st.markdown("#### Adversário provável em cada fase")
            st.caption(
                "Probabilidade de cada seleção ser o adversário, dado que o Brasil chegou àquela fase."
            )
            col_adv_1, col_adv_2, col_adv_3 = st.columns(3)
            with col_adv_1:
                st.caption("16-avos")
                show_table(detailed_tables["brasil_adversarios_16avos"], height=300)
            with col_adv_2:
                st.caption("Oitavas")
                show_table(detailed_tables["brasil_adversarios_oitavas"], height=300)
            with col_adv_3:
                st.caption("Quartas")
                show_table(detailed_tables["brasil_adversarios_quartas"], height=300)

            col_adv_4, col_adv_5, _ = st.columns(3)
            with col_adv_4:
                st.caption("Semifinal")
                show_table(detailed_tables["brasil_adversarios_semifinal"], height=300)
            with col_adv_5:
                st.caption("Final")
                show_table(detailed_tables["brasil_adversarios_final"], height=300)

            st.markdown("#### Quem pode eliminar o Brasil")
            col_elim_1, col_elim_2 = st.columns([1, 2])
            with col_elim_1:
                st.caption("Resumo — quem mais elimina")
                show_table(detailed_tables["eliminadores_brasil_agrupado"], height=320)
            with col_elim_2:
                st.caption("Por fase")
                elim_fase = detailed_tables["eliminadores_brasil"].copy()
                if "Rank" not in elim_fase.columns:
                    elim_fase.insert(0, "Rank", range(1, len(elim_fase) + 1))
                show_table(elim_fase, height=320)

            st.markdown("#### Chance de título do Brasil por condição")
            cond_df = detailed_tables["titulo_condicional_brasil"].drop(columns=["Titulos Brasil"], errors="ignore")
            show_table(cond_df, height=360)
        else:
            st.info("Simule com tabelas detalhadas ativadas para ver o caminho do Brasil.")

    with tab4:
        # ===================== 4. Finais =====================
        if detailed_tables:
            st.markdown("### 🏁 Finais mais prováveis")
            st.caption("Probabilidade de cada confronto ser a final da Copa.")
            _finais = detailed_tables["finais"][["Final", "Probabilidade"]].copy()
            _finais["Final"] = _finais["Final"].map(final_pair_label)
            _finais["Probabilidade"] = _finais["Probabilidade"] * 100
            if "Rank" not in _finais.columns:
                _finais.insert(0, "Rank", range(1, len(_finais) + 1))
            st.dataframe(
                _finais,
                use_container_width=False,
                height=360,
                hide_index=True,
                column_config={
                    "Rank": st.column_config.NumberColumn("Rank", width=50, format="%d"),
                    "Final": st.column_config.TextColumn("Final", alignment="center"),
                    "Probabilidade": pct_col(),
                },
            )
        else:
            st.info("Simule com tabelas detalhadas ativadas para ver as finais mais prováveis.")

    with tab5:
        # ===================== 5. Zebras =====================
        if detailed_tables:
            st.markdown("### 🦓 Zebras — surpresas pelo indicador de força")
            st.markdown(
                _HELP_P.format(
                    "Chance de uma surpresa: as seleções mais fracas pelo <b>indicador de força</b> "
                    "indo longe. À esquerda, as <b>16 mais fracas</b> (zebra pesada); à direita, as "
                    "<b>32 mais fracas</b> (mini-zebra, surpresa menor)."
                ),
                unsafe_allow_html=True,
            )
            col_z1, col_z2, col_z3, col_z4 = st.columns([1, 1.2, 1, 1.2])
            with col_z1:
                st.caption("16 mais fracas — quais são")
                show_table(detailed_tables["bottom16_lista"], height=250, use_container_width=True)
            with col_z2:
                st.caption("16 mais fracas — chance por fase")
                show_table(detailed_tables["bottom16_surpresa"], height=250, use_container_width=True)
            with col_z3:
                st.caption("32 mais fracas — quais são")
                show_table(detailed_tables["minizebra_lista"], height=250, use_container_width=True)
            with col_z4:
                st.caption("32 mais fracas — chance por fase")
                show_table(detailed_tables["minizebra_surpresa"], height=250, use_container_width=True)
        else:
            st.info("Simule com tabelas detalhadas ativadas para ver as zebras.")

    with tab6:
        # ===================== 6. Probabilidades por categoria =====================
        if category_tables:
            def show_category_table(df: pd.DataFrame, height: int = 420) -> None:
                keep = [c for c in ["Categoria", "Nº Seleções", "Força Média", "Campeão", "Média Campeão"] if c in df.columns]
                view = df[keep]
                prob_cols = [c for c in ["Campeão", "Média Campeão"] if c in view.columns]
                st.dataframe(
                    prep_pct(view, prob_cols),
                    use_container_width=True,
                    height=height,
                    hide_index=True,
                    column_config={
                        "Nº Seleções": st.column_config.NumberColumn(format="%d"),
                        "Força Média": st.column_config.NumberColumn(format="%.3f"),
                        "Campeão": pct_col(label="Campeão", width=None),
                        "Média Campeão": pct_col(label="Média Campeão", width=None),
                    },
                )

            st.markdown("### 📁 Probabilidades por categoria")
            st.markdown(
                _HELP_P.format(
                    "Probabilidade de o <b>campeão</b> da Copa sair de cada categoria — a soma das "
                    "chances de título das seleções do grupo. As categorias somam 100% entre si."
                ),
                unsafe_allow_html=True,
            )

            # Linha 1: Confederação e Títulos
            col_row1_left, col_row1_right = st.columns(2)
            with col_row1_left:
                st.markdown("#### Por confederação")
                show_category_table(category_tables["Por confederação"], height=280)

            with col_row1_right:
                st.markdown("#### Por títulos mundiais")
                show_category_table(category_tables["Por títulos em Copas"], height=240)

            # Linha 2: Grupo (dividido por 2 para não ir pra largura toda)
            col_row2_left, _ = st.columns(2)
            with col_row2_left:
                st.markdown("#### Por grupo")
                show_category_table(category_tables["Por grupo"], height=460)
        else:
            st.info("Simule para ver as probabilidades por categoria.")

st.markdown("---")
_sim_bytes = st.session_state["explorador_sim_excel_bytes"]
_col_dl, _ = st.columns([1, 3])
with _col_dl:
    st.download_button(
        label="Baixar Excel",
        data=_sim_bytes if _sim_bytes else "",
        file_name="simulacao_explorador_forca.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        disabled=_sim_bytes is None,
        help="Disponível após rodar uma simulação." if _sim_bytes is None else None,
    )
